"""Importación del Excel de "Audiencias".

Es el tercer reporte del PJUD que maneja el sistema y el único que mira hacia
adelante: lista las audiencias AGENDADAS de un RUT dentro de un rango de
fechas futuro (`Audiencias_{RUT}_{DD}_{MM}_{YYYY}_{DD}_{MM}_{YYYY}.xls`).

Particularidades frente al parser de movimientos:

* Cada hoja tiene un conjunto de columnas DISTINTO, no solo otro orden:
    - Familia:  Rit · Ruc · Fecha Audiencia · Tribunal · Sala · Tipo Audiencia ·
                Hora · Caratulado · Juez
    - Laboral:  Rit · Fecha audiencia · Tribunal · Sala · Tipo Audiencia · Ruc ·
                Juez · Hora · Caratulado
    - Penal:    Ruc · Fecha/Hora Inicio · Juzgado/Sala · Estado · Tipo/Audiencia ·
                Juez   (sin Rit, sin Caratulado, y con fecha y hora en una sola
                columna)
  El mapeo va por encabezado normalizado, así que las tres conviven sin
  ramificar el código.
* Los archivos consecutivos SE TRASLAPAN (el de esta semana y el de la próxima
  comparten días), así que la importación hace upsert sobre una clave natural
  en vez de insertar a ciegas. Ver `calcular_clave_natural`.
* Una fila sin fecha de audiencia parseable no sirve para nada en este módulo
  (ni se lista, ni se calendariza) y se descarta contándola aparte.
"""

import hashlib
import logging
import os
import re
import unicodedata
from datetime import date, datetime, time, timezone
from typing import Optional

import openpyxl
import xlrd
from sqlalchemy.orm import Session

from app.models.audiencia import Audiencia
from app.models.estado_diario_origen import EstadoDiarioOrigen
from app.repositories.audiencia_repository import AudienciaRepository
from app.repositories.jurisdiccion_repository import JurisdiccionRepository
from app.utils.excel_pjud import (
    detectar_formato,
    mapa_columnas,
    parse_fecha_xls,
    parse_fecha_xlsx,
    parse_hora_xls,
    parse_hora_xlsx,
    recortar,
)

logger = logging.getLogger(__name__)

# Encabezado normalizado -> campo del modelo Audiencia.
# "Rit" es el rol de primera instancia, igual que en el módulo Movimientos.
HEADER_ALIASES = {
    "rit": "rol",
    "rol": "rol",
    "ruc": "ruc",
    "fecha audiencia": "fecha_audiencia",
    "fecha de audiencia": "fecha_audiencia",
    "tribunal": "tribunal",
    # La hoja Penal junta juzgado y sala en una sola columna. Se guarda entera
    # como tribunal: partirla por "/" rompería nombres que ya lo contienen
    # ("Jgdo. L. de Tocopilla / Sala 1" vs "Juzgado de Garantía S/N").
    "juzgado/sala": "tribunal",
    "sala": "sala",
    "tipo audiencia": "tipo_audiencia",
    "tipo/audiencia": "tipo_audiencia",
    "tipo de audiencia": "tipo_audiencia",
    "hora": "hora",
    "caratulado": "caratulado",
    "juez": "juez",
    "estado": "estado",
    # Penal: fecha y hora en una sola celda. Se parsea a las dos columnas.
    "fecha/hora inicio": "fecha_hora_inicio",
    "fecha/hora": "fecha_hora_inicio",
}

CAMPOS_FECHA = ("fecha_audiencia",)
CAMPOS_HORA = ("hora",)
# Se resuelve aparte: alimenta fecha_audiencia y hora a la vez.
CAMPO_FECHA_HORA = "fecha_hora_inicio"

LARGOS_MAXIMOS = {
    "materia": 100,
    "rol": 100,
    "ruc": 50,
    "caratulado": 255,
    "tribunal": 255,
    "sala": 100,
    "tipo_audiencia": 255,
    "juez": 255,
    "estado": 100,
}

# Campos que identifican la audiencia. `sala`, `juez` y `estado` quedan FUERA a
# propósito: cambian entre un archivo y el siguiente (recalendarizaciones,
# cambio de juez) y si entraran en la clave la misma audiencia se duplicaría en
# vez de actualizarse.
CAMPOS_CLAVE = ("materia", "rol", "ruc", "tribunal", "tipo_audiencia")

# Campos que sí se refrescan cuando la audiencia vuelve a llegar.
CAMPOS_ACTUALIZABLES = ("sala", "juez", "estado", "caratulado")


def _normalizar_para_clave(valor) -> str:
    """Minúsculas sin tildes ni espacios repetidos.

    El PJUD manda el mismo tribunal con distinto relleno de espacios entre un
    archivo y otro; sin normalizar, la deduplicación no engancharía.
    """
    if valor in (None, ""):
        return ""
    texto = unicodedata.normalize("NFKD", str(valor))
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", texto).strip().lower()


def calcular_clave_natural(fila: dict) -> str:
    """Identidad de una audiencia, estable entre archivos traslapados.

    Es un SHA-1 y no las columnas sueltas porque casi todas son NULL-ables y en
    PostgreSQL dos NULL nunca colisionan: un UNIQUE sobre las columnas dejaría
    pasar duplicados justo en las filas peor pobladas (la hoja Penal, que no
    trae Rit).
    """
    partes = [_normalizar_para_clave(fila.get(c)) for c in CAMPOS_CLAVE]
    fecha = fila.get("fecha_audiencia")
    hora = fila.get("hora")
    partes.append(fecha.isoformat() if isinstance(fecha, date) else "")
    partes.append(hora.isoformat() if isinstance(hora, time) else "")
    return hashlib.sha1("|".join(partes).encode("utf-8")).hexdigest()


def _recortar(valor: Optional[str], campo: str) -> Optional[str]:
    return recortar(valor, campo, LARGOS_MAXIMOS)


def _mapa_columnas(encabezados: list) -> dict:
    return mapa_columnas(encabezados, HEADER_ALIASES)


def _armar_fila(valores: dict, materia: str) -> Optional[dict]:
    """Fila lista para persistir, o None si venía vacía o sin fecha.

    Sin `fecha_audiencia` la fila no se puede listar ni calendarizar, que es
    todo lo que hace este módulo: se descarta en vez de guardarse muda.
    """
    if not any(v not in (None, "") for v in valores.values()):
        return None
    if not isinstance(valores.get("fecha_audiencia"), date):
        return None

    fila = {"materia": _recortar(materia.strip(), "materia")}
    fila.update(valores)
    fila["clave_natural"] = calcular_clave_natural(fila)
    return fila


def _valor_texto(bruto, campo: str) -> Optional[str]:
    texto = str(bruto).strip() if bruto not in (None, "") else None
    return _recortar(texto or None, campo)


def _leer_xls(file_path: str) -> tuple[list[dict], int]:
    """Devuelve (filas, descartadas_sin_fecha)."""
    wb = xlrd.open_workbook(file_path)
    filas: list[dict] = []
    sin_fecha = 0

    for hoja in wb.sheets():
        if hoja.nrows < 2:  # solo encabezado (o vacía)
            continue

        mapa = _mapa_columnas([hoja.cell_value(0, c) for c in range(hoja.ncols)])
        if not mapa:
            continue

        for r in range(1, hoja.nrows):
            valores: dict = {}
            for col, campo in mapa.items():
                bruto = hoja.cell_value(r, col)
                if campo == CAMPO_FECHA_HORA:
                    valores["fecha_audiencia"] = parse_fecha_xls(bruto, wb.datemode)
                    valores["hora"] = parse_hora_xls(bruto, wb.datemode)
                elif campo in CAMPOS_FECHA:
                    valores[campo] = parse_fecha_xls(bruto, wb.datemode)
                elif campo in CAMPOS_HORA:
                    valores[campo] = parse_hora_xls(bruto, wb.datemode)
                else:
                    valores[campo] = _valor_texto(bruto, campo)

            fila = _armar_fila(valores, hoja.name)
            if fila:
                filas.append(fila)
            elif any(v not in (None, "") for v in valores.values()):
                sin_fecha += 1

    return filas, sin_fecha


def _leer_xlsx(file_path: str) -> tuple[list[dict], int]:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    try:
        filas: list[dict] = []
        sin_fecha = 0

        for nombre_hoja in wb.sheetnames:
            hoja = wb[nombre_hoja]
            if not hoja.max_row or hoja.max_row < 2:
                continue

            ncols = hoja.max_column or 0
            mapa = _mapa_columnas([hoja.cell(1, c).value for c in range(1, ncols + 1)])
            if not mapa:
                continue

            for r in range(2, hoja.max_row + 1):
                valores: dict = {}
                for col, campo in mapa.items():
                    bruto = hoja.cell(r, col + 1).value  # openpyxl es 1-based
                    if campo == CAMPO_FECHA_HORA:
                        valores["fecha_audiencia"] = parse_fecha_xlsx(bruto)
                        valores["hora"] = parse_hora_xlsx(bruto)
                    elif campo in CAMPOS_FECHA:
                        valores[campo] = parse_fecha_xlsx(bruto)
                    elif campo in CAMPOS_HORA:
                        valores[campo] = parse_hora_xlsx(bruto)
                    else:
                        valores[campo] = _valor_texto(bruto, campo)

                fila = _armar_fila(valores, nombre_hoja)
                if fila:
                    filas.append(fila)
                elif any(v not in (None, "") for v in valores.values()):
                    sin_fecha += 1

        return filas, sin_fecha
    finally:
        wb.close()


def parse_audiencias_file(file_path: str) -> tuple[list[dict], int]:
    """Parsea el Excel de audiencias a (filas, descartadas_sin_fecha).

    Función pura (no toca la base de datos) para poder probar el parseo del
    archivo por separado.
    """
    if detectar_formato(file_path) == "xlsx":
        return _leer_xlsx(file_path)
    return _leer_xls(file_path)


_PATRON_NOMBRE = re.compile(
    r"^Audiencias_(\d+(?:[\-]?[0-9kK])?)_+"
    r"(\d{1,2})_(\d{1,2})_(\d{4})_+"
    r"(\d{1,2})_(\d{1,2})_(\d{4})$",
    re.IGNORECASE,
)


def parse_nombre_archivo(
    filename: str,
) -> tuple[Optional[str], Optional[date], Optional[date]]:
    """Extrae (RUT, fecha_desde, fecha_hasta) del nombre del archivo.

    Formato real: `Audiencias_16952077_03_08_2026_09_08_2026.xls`, o sea RUT
    seguido del rango de fechas que cubre el reporte. Se usa `_+` como
    separador porque el PJUD dobla el guion bajo cuando el RUT viene sin
    dígito verificador, igual que en el archivo de movimientos.
    """
    nombre = os.path.splitext(filename or "")[0]
    m = _PATRON_NOMBRE.match(nombre)
    if not m:
        return None, None, None

    rut = m.group(1)
    try:
        desde = date(int(m.group(4)), int(m.group(3)), int(m.group(2)))
    except ValueError:
        desde = None
    try:
        hasta = date(int(m.group(7)), int(m.group(6)), int(m.group(5)))
    except ValueError:
        hasta = None
    return rut, desde, hasta


class AudienciaImportService:
    """Importa un archivo de audiencias deduplicando contra lo ya cargado.

    El contrato de `import_file` es el mismo que el de ImportService y
    MovimientoImportService (mismos parámetros, y `movimientos_importados` en el
    resultado) para que CorreoService pueda tratar los tres tipos por igual.
    """

    def __init__(self, db: Session):
        self.db = db
        self.repo = AudienciaRepository(db)
        self.jurisdiccion_repo = JurisdiccionRepository(db)

    def import_file(
        self,
        file_path: str,
        rut: str,
        fecha: date,
        usuario_id: Optional[int] = None,
        nombre_archivo: Optional[str] = None,
    ) -> dict:
        if usuario_id is None:
            # El UNIQUE de deduplicación es (usuario_id, clave_natural) y
            # usuario_id es NOT NULL: sin dueño no hay dónde guardar la fila.
            raise ValueError(
                "No se puede importar audiencias sin un usuario dueño del archivo"
            )

        # Duplicado a nivel de ARCHIVO, medido dentro del tipo: el mismo RUT
        # puede tener el mismo día un estado diario, un archivo de movimientos y
        # uno de audiencias sin que ninguno sea duplicado del otro.
        existente = self.repo.find_origen_audiencias(rut, fecha)
        if existente:
            raise ValueError(
                f"Ya existe un archivo de audiencias para el RUT {rut} del {fecha} "
                f"(origen {existente.id}). Elimínelo antes de volver a cargarlo."
            )

        try:
            filas, sin_fecha = parse_audiencias_file(file_path)
        except Exception as e:
            raise ValueError(f"No se pudo leer el archivo de audiencias: {e}")

        if not filas:
            raise ValueError(
                "El archivo no contiene audiencias reconocibles. Verifique que las "
                "hojas tengan los encabezados esperados (Rit/Ruc, Fecha Audiencia, "
                "Tribunal, Sala, Tipo Audiencia, Hora, Caratulado, Juez)."
            )

        origen = EstadoDiarioOrigen(
            usuario_carga_id=usuario_id,
            tipo=EstadoDiarioOrigen.TIPO_AUDIENCIAS,
            rut=rut,
            fecha=fecha,
            nombre_archivo=nombre_archivo or os.path.basename(file_path),
            fecha_carga=datetime.now(timezone.utc),
        )
        self.db.add(origen)
        self.db.flush()  # asigna origen.id sin cerrar la transacción

        nuevas, actualizadas = self._persistir(filas, origen, usuario_id)
        self.db.commit()

        por_materia: dict[str, int] = {}
        for fila in filas:
            clave = fila.get("materia") or "(sin materia)"
            por_materia[clave] = por_materia.get(clave, 0) + 1

        logger.info(
            "Audiencias del origen %d: %d nuevas, %d actualizadas, %d sin fecha (%s)",
            origen.id, nuevas, actualizadas, sin_fecha, por_materia,
        )

        # Google Calendar va DESPUÉS del commit y es best-effort: las audiencias
        # ya están guardadas y ningún flujo depende de que Google responda.
        self._sincronizar_google(usuario_id)

        return {
            "origen_id": origen.id,
            # Nombre heredado del contrato común de importación; acá son las
            # audiencias que el archivo aportó realmente.
            "movimientos_importados": nuevas + actualizadas,
            "audiencias_nuevas": nuevas,
            "audiencias_actualizadas": actualizadas,
            "audiencias_sin_fecha": sin_fecha,
            "por_materia": por_materia,
        }

    def _persistir(
        self, filas: list[dict], origen: EstadoDiarioOrigen, usuario_id: int
    ) -> tuple[int, int]:
        """Inserta las audiencias nuevas y refresca las ya conocidas.

        Las existentes se traen de una sola consulta por clave (no una por fila)
        para no caer en N+1 con archivos de cientos de audiencias.
        """
        # Un mismo archivo puede repetir la clave entre hojas; la última gana.
        por_clave = {fila["clave_natural"]: fila for fila in filas}
        existentes = self.repo.find_por_claves(usuario_id, list(por_clave.keys()))

        cache_jurisdiccion: dict[str, Optional[int]] = {}
        nuevas = 0

        for clave, fila in por_clave.items():
            audiencia = existentes.get(clave)
            if audiencia is None:
                audiencia = Audiencia(
                    usuario_id=usuario_id,
                    clave_natural=clave,
                    materia=fila.get("materia"),
                    rol=fila.get("rol"),
                    ruc=fila.get("ruc"),
                    tribunal=fila.get("tribunal"),
                    tipo_audiencia=fila.get("tipo_audiencia"),
                    fecha_audiencia=fila["fecha_audiencia"],
                    hora=fila.get("hora"),
                )
                self.db.add(audiencia)
                nuevas += 1

            # Procedencia: siempre el archivo más reciente que la trajo.
            audiencia.estado_diario_origen_id = origen.id
            for campo in CAMPOS_ACTUALIZABLES:
                setattr(audiencia, campo, fila.get(campo))

            # La materia es el nombre de la hoja (Familia, Laboral, Penal) y
            # calza con el catálogo de jurisdicciones sembrado al inicio.
            if audiencia.materia:
                if audiencia.materia not in cache_jurisdiccion:
                    jur = self.jurisdiccion_repo.find_by_nombre(audiencia.materia)
                    cache_jurisdiccion[audiencia.materia] = jur.id if jur else None
                audiencia.jurisdiccion_id = cache_jurisdiccion[audiencia.materia]

        return nuevas, len(por_clave) - nuevas

    def _sincronizar_google(self, usuario_id: int) -> None:
        """Empuja al Google Calendar del dueño las audiencias futuras.

        Import local: GoogleCalendarService arrastra el cliente de Google y este
        módulo se importa desde el parser, que debe poder correr sin esa
        dependencia. Nunca propaga excepciones: la importación ya está commiteada.
        """
        try:
            from app.services.audiencia_calendar_service import AudienciaCalendarService

            AudienciaCalendarService(self.db).sincronizar_pendientes(usuario_id)
        except Exception:
            logger.exception(
                "No se pudieron sincronizar con Google las audiencias del usuario %s",
                usuario_id,
            )
