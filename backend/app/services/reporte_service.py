"""Informes dinámicos: el usuario elige qué campos quiere, se arma un Excel y
se le manda por correo.

El catálogo de campos (`CAMPOS`) es la única fuente de verdad: define qué se
puede pedir, cómo se rotula en el Excel y cómo se extrae del modelo. La UI lo
consulta para pintar el selector, y la generación valida contra él. Así una
plantilla guardada no puede pedir un campo que ya no existe, y agregar un campo
nuevo es una línea acá y nada más.
"""

import io
import logging
from datetime import date, datetime
from typing import Callable, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, joinedload

from app.core.database import SesionMaestra
from app.core.exceptions import BadRequestException, NotFoundException
from app.models.estado_diario import EstadoDiario
from app.models.estado_diario_agenda import EstadoDiarioAgenda
from app.models.estado_diario_origen import EstadoDiarioOrigen
from app.models.movimiento import Movimiento
from app.models.reporte_plantilla import ReportePlantilla
from app.services.smtp_service import SmtpService

logger = logging.getLogger(__name__)

# Tope de filas por informe. Un Excel de 50 mil filas no lo abre nadie y el
# adjunto no pasa los límites de tamaño de la mayoría de los servidores.
MAX_FILAS = 20000


def _fecha(valor) -> Optional[str]:
    """Formato chileno. Las fechas van como texto para que Excel no las
    reinterprete según la configuración regional de quien abre el archivo."""
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor.strftime("%d-%m-%Y %H:%M")
    if isinstance(valor, date):
        return valor.strftime("%d-%m-%Y")
    return str(valor)


def _si_no(valor) -> str:
    return "Sí" if valor else "No"


class Campo:
    __slots__ = ("clave", "etiqueta", "extraer")

    def __init__(self, clave: str, etiqueta: str, extraer: Callable):
        self.clave = clave
        self.etiqueta = etiqueta
        self.extraer = extraer


def _c(clave: str, etiqueta: str, extraer: Callable) -> tuple[str, Campo]:
    return clave, Campo(clave, etiqueta, extraer)


# ── Catálogo de campos por fuente ────────────────────────────────────────

CAMPOS: dict[str, dict[str, Campo]] = {
    ReportePlantilla.FUENTE_ESTADO_DIARIO: dict([
        _c("rol", "Rol", lambda m: m.rol),
        _c("rol_unico", "Rol único", lambda m: m.rol_unico),
        _c("caratulado", "Caratulado", lambda m: m.caratulado),
        _c("tribunal", "Tribunal", lambda m: m.tribunal),
        _c("corte", "Corte", lambda m: m.corte),
        _c("jurisdiccion", "Jurisdicción",
           lambda m: m.jurisdiccion.nombre if m.jurisdiccion else None),
        _c("tipo_causa", "Tipo de causa", lambda m: m.tipo_causa),
        _c("estado", "Estado de la causa", lambda m: m.estado),
        _c("ubicacion", "Ubicación", lambda m: m.ubicacion),
        _c("fecha_ingreso", "Fecha de ingreso", lambda m: _fecha(m.fecha_ingreso)),
        _c("fecha_ubicacion", "Fecha de ubicación", lambda m: _fecha(m.fecha_ubicacion)),
        _c("rut", "RUT",
           lambda m: m.estado_diario_origen.rut if m.estado_diario_origen else None),
        _c("fecha_estado_diario", "Fecha del estado diario",
           lambda m: _fecha(m.estado_diario_origen.fecha) if m.estado_diario_origen else None),
        _c("nombre_archivo", "Archivo de origen",
           lambda m: m.estado_diario_origen.nombre_archivo if m.estado_diario_origen else None),
        _c("resuelto", "Resuelto", lambda m: _si_no(m.leido)),
        _c("fecha_leido", "Fecha de resolución", lambda m: _fecha(m.fecha_leido)),
        _c("observacion_resuelto", "Observación al resolver",
           lambda m: m.observacion_resuelto),
        _c("pendiente", "Pendiente", lambda m: _si_no(m.pendiente)),
        _c("nivel_pendiente", "Nivel de urgencia", lambda m: m.nivel_pendiente),
        _c("fecha_pendiente", "Fecha en que quedó pendiente",
           lambda m: _fecha(m.fecha_pendiente)),
        _c("usuario_pendiente", "Responsable",
           lambda m: m.usuario_pendiente.usuario if m.usuario_pendiente else None),
    ]),
    ReportePlantilla.FUENTE_MOVIMIENTOS: dict([
        _c("materia", "Materia", lambda m: m.materia),
        _c("rol", "Rol / RIT", lambda m: m.rol),
        _c("era", "Era", lambda m: m.era),
        _c("caratulado", "Caratulado", lambda m: m.caratulado),
        _c("tribunal", "Tribunal", lambda m: m.tribunal),
        _c("corte", "Corte", lambda m: m.corte),
        _c("fecha_ingreso", "Fecha de ingreso", lambda m: _fecha(m.fecha_ingreso)),
        _c("estado_causa", "Estado de la causa", lambda m: m.estado_causa),
        _c("institucion", "Institución", lambda m: m.institucion),
        _c("ubicacion", "Ubicación", lambda m: m.ubicacion),
        _c("fecha_ubicacion", "Fecha de ubicación", lambda m: _fecha(m.fecha_ubicacion)),
        _c("rut", "RUT",
           lambda m: m.estado_diario_origen.rut if m.estado_diario_origen else None),
        _c("fecha_archivo", "Fecha del archivo",
           lambda m: _fecha(m.estado_diario_origen.fecha) if m.estado_diario_origen else None),
        _c("nombre_archivo", "Archivo de origen",
           lambda m: m.estado_diario_origen.nombre_archivo if m.estado_diario_origen else None),
    ]),
    ReportePlantilla.FUENTE_AGENDA: dict([
        _c("detalle", "Detalle del recordatorio", lambda a: a.detalle),
        _c("fecha_hora", "Fecha del recordatorio", lambda a: _fecha(a.fecha_hora)),
        _c("nivel", "Nivel de urgencia", lambda a: a.nivel),
        _c("finalizado", "Finalizado", lambda a: _si_no(a.finalizado)),
        _c("fecha_finalizacion", "Fecha de finalización",
           lambda a: _fecha(a.fecha_finalizacion)),
        _c("usuario_registro", "Creado por",
           lambda a: a.usuario_registro.usuario if a.usuario_registro else None),
        _c("notificar_whatsapp", "Notifica por WhatsApp",
           lambda a: _si_no(a.notificar_whatsapp)),
        _c("enviado", "WhatsApp enviado", lambda a: _si_no(a.enviado)),
        _c("caratulado", "Caratulado",
           lambda a: a.estado_diario.caratulado if a.estado_diario else None),
        _c("rol", "Rol", lambda a: a.estado_diario.rol if a.estado_diario else None),
        _c("tribunal", "Tribunal",
           lambda a: a.estado_diario.tribunal if a.estado_diario else None),
    ]),
}

ETIQUETA_FUENTE = {
    ReportePlantilla.FUENTE_ESTADO_DIARIO: "Estado Diario",
    ReportePlantilla.FUENTE_MOVIMIENTOS: "Movimientos",
    ReportePlantilla.FUENTE_AGENDA: "Recordatorios",
}


class ReporteService:
    def __init__(self, db: Session):
        self.db = db

    # ── Catálogo ──────────────────────────────────────────

    @staticmethod
    def campos_disponibles() -> list[dict]:
        """Lo que la UI pinta en el selector de campos."""
        return [
            {
                "fuente": fuente,
                "etiqueta": ETIQUETA_FUENTE[fuente],
                "campos": [
                    {"clave": c.clave, "etiqueta": c.etiqueta} for c in campos.values()
                ],
            }
            for fuente, campos in CAMPOS.items()
        ]

    @staticmethod
    def validar(fuente: str, campos: list[str]) -> list[str]:
        """Deja solo los campos que existen en la fuente, en el orden que eligió
        el usuario (ese orden ES el orden de columnas del Excel)."""
        if fuente not in CAMPOS:
            raise BadRequestException(
                f"Fuente de datos desconocida: {fuente}. "
                f"Válidas: {', '.join(CAMPOS)}"
            )
        catalogo = CAMPOS[fuente]
        validos = [c for c in campos if c in catalogo]
        if not validos:
            raise BadRequestException(
                "Debe elegir al menos un campo válido para el informe"
            )
        return validos

    # ── Consulta ──────────────────────────────────────────

    def _consultar(self, fuente: str, filtros: dict, scope_usuario_id: Optional[int]):
        """Filas del informe, ya acotadas al dueño.

        `scope_usuario_id=None` = admin (sin filtro), misma convención que el
        resto del sistema.
        """
        if fuente == ReportePlantilla.FUENTE_MOVIMIENTOS:
            query = (
                self.db.query(Movimiento)
                .join(Movimiento.estado_diario_origen)
                .options(joinedload(Movimiento.estado_diario_origen))
            )
            if scope_usuario_id is not None:
                query = query.filter(EstadoDiarioOrigen.usuario_carga_id == scope_usuario_id)
            if filtros.get("materia"):
                query = query.filter(Movimiento.materia == filtros["materia"])
            if filtros.get("estado_causa"):
                query = query.filter(Movimiento.estado_causa == filtros["estado_causa"])
            query = self._filtrar_fechas_archivo(query, filtros)
            return query.order_by(Movimiento.id).limit(MAX_FILAS).all()

        if fuente == ReportePlantilla.FUENTE_AGENDA:
            query = (
                self.db.query(EstadoDiarioAgenda)
                .options(
                    joinedload(EstadoDiarioAgenda.estado_diario),
                    joinedload(EstadoDiarioAgenda.usuario_registro),
                )
            )
            if scope_usuario_id is not None:
                query = query.filter(
                    EstadoDiarioAgenda.usuario_registro_id == scope_usuario_id
                )
            estado = filtros.get("estado")
            if estado == "vigentes":
                query = query.filter(EstadoDiarioAgenda.finalizado.is_(False))
            elif estado == "finalizados":
                query = query.filter(EstadoDiarioAgenda.finalizado.is_(True))
            if filtros.get("nivel"):
                query = query.filter(EstadoDiarioAgenda.nivel == filtros["nivel"])
            if filtros.get("fecha_desde"):
                query = query.filter(EstadoDiarioAgenda.fecha_hora >= filtros["fecha_desde"])
            if filtros.get("fecha_hasta"):
                query = query.filter(EstadoDiarioAgenda.fecha_hora <= filtros["fecha_hasta"])
            return query.order_by(EstadoDiarioAgenda.fecha_hora).limit(MAX_FILAS).all()

        # Estado diario (por defecto)
        query = (
            self.db.query(EstadoDiario)
            .join(EstadoDiario.estado_diario_origen)
            .options(
                joinedload(EstadoDiario.estado_diario_origen),
                joinedload(EstadoDiario.jurisdiccion),
                joinedload(EstadoDiario.usuario_pendiente),
            )
        )
        if scope_usuario_id is not None:
            query = query.filter(EstadoDiarioOrigen.usuario_carga_id == scope_usuario_id)

        estado = filtros.get("estado")
        if estado == "resuelto":
            query = query.filter(EstadoDiario.leido.is_(True))
        elif estado == "pendiente":
            query = query.filter(
                EstadoDiario.pendiente.is_(True), EstadoDiario.leido.is_(False)
            )
        elif estado == "no-leido":
            query = query.filter(
                EstadoDiario.leido.is_(False), EstadoDiario.pendiente.is_(False)
            )

        if filtros.get("nivel"):
            query = query.filter(EstadoDiario.nivel_pendiente == filtros["nivel"])
        if filtros.get("jurisdiccion_id"):
            query = query.filter(EstadoDiario.jurisdiccion_id == filtros["jurisdiccion_id"])
        query = self._filtrar_fechas_archivo(query, filtros)
        return query.order_by(EstadoDiario.id).limit(MAX_FILAS).all()

    @staticmethod
    def _filtrar_fechas_archivo(query, filtros: dict):
        """El rango de fechas se aplica sobre la fecha del ARCHIVO, no sobre la
        fecha de ingreso de la causa: el usuario piensa en "el estado diario de
        esta semana", no en cuándo entró la causa al tribunal."""
        if filtros.get("fecha_desde"):
            query = query.filter(EstadoDiarioOrigen.fecha >= filtros["fecha_desde"])
        if filtros.get("fecha_hasta"):
            query = query.filter(EstadoDiarioOrigen.fecha <= filtros["fecha_hasta"])
        return query

    # ── Excel ─────────────────────────────────────────────

    def generar_excel(self, plantilla: ReportePlantilla, scope_usuario_id: Optional[int]) -> bytes:
        campos = self.validar(plantilla.fuente, plantilla.lista_campos)
        catalogo = CAMPOS[plantilla.fuente]
        filas = self._consultar(plantilla.fuente, plantilla.dict_filtros, scope_usuario_id)

        wb = Workbook()
        ws = wb.active
        ws.title = ETIQUETA_FUENTE[plantilla.fuente][:31]  # Excel corta en 31

        encabezado_fondo = PatternFill("solid", fgColor="1D4ED8")  # primary-700
        encabezado_fuente = Font(color="FFFFFF", bold=True)

        for col, clave in enumerate(campos, start=1):
            celda = ws.cell(row=1, column=col, value=catalogo[clave].etiqueta)
            celda.fill = encabezado_fondo
            celda.font = encabezado_fuente
            celda.alignment = Alignment(vertical="center")

        for fila_idx, registro in enumerate(filas, start=2):
            for col, clave in enumerate(campos, start=1):
                try:
                    valor = catalogo[clave].extraer(registro)
                except Exception:
                    # Un campo que falla no puede botar el informe completo.
                    logger.warning("No se pudo extraer '%s' de la fila %s", clave, registro)
                    valor = None
                ws.cell(row=fila_idx, column=col, value=valor)

        # Ancho por contenido, con tope: una carátula larga no debe dejar una
        # columna de 200 caracteres.
        for col, clave in enumerate(campos, start=1):
            largo = len(catalogo[clave].etiqueta)
            for fila_idx in range(2, min(len(filas) + 2, 200)):
                valor = ws.cell(row=fila_idx, column=col).value
                if valor is not None:
                    largo = max(largo, len(str(valor)))
            ws.column_dimensions[get_column_letter(col)].width = min(max(largo + 2, 10), 50)

        ws.freeze_panes = "A2"
        if filas:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(campos))}{len(filas) + 1}"

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    # ── Generar y enviar ──────────────────────────────────

    def generar_y_enviar(self, plantilla_id: int, current_user) -> dict:
        from app.repositories.reporte_repository import ReporteRepository

        scope = None if current_user.rol == "admin" else current_user.id
        plantilla = ReporteRepository(self.db).find_by_id(plantilla_id, scope)
        if plantilla is None:
            raise NotFoundException("Informe no encontrado")

        if not current_user.correo:
            raise BadRequestException(
                "Su usuario no tiene un correo registrado. Agréguelo en su perfil "
                "para poder recibir informes."
            )

        contenido = self.generar_excel(plantilla, scope)

        hoy = datetime.now().strftime("%d-%m-%Y")
        nombre_archivo = f"{_slug(plantilla.nombre)}_{datetime.now():%Y%m%d_%H%M}.xlsx"
        cuerpo = (
            f"Estimado/a {current_user.nombre or current_user.usuario}:\n\n"
            f"Adjunto encontrará el informe «{plantilla.nombre}» generado el {hoy}.\n\n"
            f"Fuente de datos: {ETIQUETA_FUENTE[plantilla.fuente]}\n"
            f"Campos incluidos: {', '.join(plantilla.lista_campos)}\n\n"
            "Este correo fue generado automáticamente por el sistema Estado Diario.\n"
        )

        # La cuenta de envío vive en la base PRINCIPAL, no en la del cliente:
        # se abre una sesión corta solo para eso y se cierra al terminar.
        db_maestra = SesionMaestra()
        try:
            SmtpService(db_maestra).enviar_con_adjunto(
                destinatario=current_user.correo,
                asunto=f"Informe: {plantilla.nombre} ({hoy})",
                cuerpo=cuerpo,
                adjunto=contenido,
                nombre_adjunto=nombre_archivo,
            )
        except Exception as e:
            plantilla.ultima_generacion = datetime.now()
            plantilla.ultimo_resultado = f"Error: {e}"
            self.db.commit()
            raise BadRequestException(str(e))
        finally:
            db_maestra.close()

        plantilla.ultima_generacion = datetime.now()
        plantilla.ultimo_resultado = f"Enviado a {current_user.correo}"
        self.db.commit()

        return {
            "exito": True,
            "mensaje": f"El informe fue enviado a {current_user.correo}",
            "archivo": nombre_archivo,
        }


def _slug(texto: str) -> str:
    """Nombre de archivo seguro: sin espacios, acentos ni separadores de ruta."""
    import re
    import unicodedata

    normalizado = unicodedata.normalize("NFKD", texto or "informe")
    sin_tildes = normalizado.encode("ascii", "ignore").decode()
    limpio = re.sub(r"[^A-Za-z0-9]+", "_", sin_tildes).strip("_")
    return (limpio or "informe")[:60]
