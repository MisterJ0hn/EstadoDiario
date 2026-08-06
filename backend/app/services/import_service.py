import logging
import os
import shutil
import unicodedata
from datetime import datetime, timezone, date
from typing import Optional

import xlrd
import openpyxl
from sqlalchemy.orm import Session

from app.models.estado_diario import EstadoDiario
from app.models.estado_diario_corte import EstadoDiarioCorte
from app.models.estado_diario_origen import EstadoDiarioOrigen
from app.repositories.estado_diario_repository import EstadoDiarioRepository
from app.repositories.estado_diario_origen_repository import EstadoDiarioOrigenRepository
from app.repositories.jurisdiccion_repository import JurisdiccionRepository
from app.models.jurisdiccion import Jurisdiccion
from app.services.deteccion_archivo import verificar_contenido

logger = logging.getLogger(__name__)

# Mapeo de nombres de columna del Excel a campos internos
# Cada hoja puede tener columnas distintas; se busca por nombre de header
# Hojas que NO son de materia sino de corte. Sus filas van a
# `estado_diario_corte`, no a `estado_diario`: no traen rol único, ni tribunal,
# ni estado de la causa, y mezclarlas dejaba media tabla vacía y el listado por
# materia mostrando filas que no calzaban con sus columnas.
#
# Se comparan sin acentos y en minúsculas porque el nombre de la hoja cambia
# entre archivos ("Corte Apelaciones", "Corte de Apelaciones").
HOJAS_CORTE = {
    "corte suprema": "suprema",
    "corte apelaciones": "apelaciones",
    "corte de apelaciones": "apelaciones",
}

# Encabezados propios de las hojas de corte. Van aparte de HEADER_ALIASES
# porque chocan: "tipo recurso" es `tipo_causa` en una hoja de materia y
# `tipo_recurso` acá, y el número de ingreso no es el `rol` de una causa.
HEADER_ALIASES_CORTE = {
    "n ingreso": "numero_ingreso",
    "n de ingreso": "numero_ingreso",
    "numero ingreso": "numero_ingreso",
    "numero de ingreso": "numero_ingreso",
    "rol": "numero_ingreso",
    "fecha ingreso": "fecha_ingreso",
    "caratulado": "caratulado",
    "ubicacion": "ubicacion",
    "fecha ubicacion": "fecha_ubicacion",
    "corte": "corte",
    "tipo recurso": "tipo_recurso",
}

HEADER_ALIASES = {
    "rol": "rol",
    "n° ingreso": "rol",
    "nø ingreso": "rol",
    "n\u00b0 ingreso": "rol",
    "rol interno": "rol",
    "rit": "rol",
    "ruc": "rol_unico",
    "rol unico": "rol_unico",
    "rol único": "rol_unico",
    "rol \u00fanico": "rol_unico",
    "fecha ingreso": "fecha_ingreso",
    "caratulado": "caratulado",
    "tribunal": "tribunal",
    "tipo causa": "tipo_causa",
    "tipo recurso": "tipo_causa",
    "ubicacion": "ubicacion",
    "ubicación": "ubicacion",
    "fecha ubicacion": "fecha_ubicacion",
    "fecha ubicación": "fecha_ubicacion",
    "estado": "estado",
    "corte": "corte",
}


def normalizar_texto(valor: str) -> str:
    """Minúsculas, sin acentos y con los espacios colapsados.

    Los encabezados y los nombres de hoja llegan escritos de varias formas
    entre un archivo y otro ("N° Ingreso", "Número de Ingreso", "Ubicación"),
    así que compararlos tal cual falla en cuanto cambia el archivo.
    """
    if not valor:
        return ""
    sin_acentos = "".join(
        c for c in unicodedata.normalize("NFD", str(valor))
        if unicodedata.category(c) != "Mn"
    )
    limpio = sin_acentos.replace("°", "").replace("º", "").replace(".", " ")
    return " ".join(limpio.lower().split())


def tipo_de_hoja_corte(nombre_hoja: str) -> Optional[str]:
    """Devuelve 'suprema'/'apelaciones' si la hoja es de corte; si no, None."""
    return HOJAS_CORTE.get(normalizar_texto(nombre_hoja))


class ImportService:
    # El estado diario es el reporte de UN día y esa fecha no está dentro del
    # archivo: si el nombre no la trae, no hay de dónde sacarla.
    deduce_fecha_del_contenido = False
    # Su control de duplicados es (rut, fecha): sin RUT dejaría entrar el mismo
    # archivo cuantas veces llegue.
    requiere_rut = True

    def __init__(self, db: Session):
        self.db = db
        self.origen_repo = EstadoDiarioOrigenRepository(db)
        self.ed_repo = EstadoDiarioRepository(db)
        self.jurisdiccion_repo = JurisdiccionRepository(db)

    def import_file(
        self,
        file_path: str,
        rut: str,
        fecha: date,
        usuario_id: Optional[int] = None,
        nombre_archivo: Optional[str] = None,
    ) -> dict:
        # Este parser acepta las columnas del reporte de movimientos casi
        # enteras (Rit, Tribunal, Caratulado, Fecha Ingreso), así que un archivo
        # de movimientos mal ruteado entraría sin dar error y quedaría grabado
        # como estado diario. Por eso el contenido se verifica antes de tocar
        # nada: es lo único que distingue los dos reportes con certeza.
        error = verificar_contenido(file_path, EstadoDiarioOrigen.TIPO_ESTADO_DIARIO)
        if error:
            raise ValueError(error)

        # Un mismo RUT solo puede tener un estado diario por fecha. Sin esta
        # validación, reimportar el archivo duplica todos los movimientos, y
        # con la ingesta automática por correo eso ocurriría sin que nadie lo
        # note. Para recargar, primero hay que borrar el origen existente.
        # La unicidad es de todo el estudio: la base es de un solo cliente y la
        # casilla de ingesta es una, así que el mismo RUT y fecha es el mismo
        # archivo, lo suba quien lo suba.
        existente = self.origen_repo.find_by_rut_fecha(
            rut, fecha, EstadoDiarioOrigen.TIPO_ESTADO_DIARIO
        )
        if existente:
            raise ValueError(
                f"Ya existe un estado diario para el RUT {rut} del {fecha} "
                f"(origen {existente.id}). Elimínelo antes de volver a cargarlo."
            )

        # Detectar formato real por header bytes (la extensión puede mentir)
        file_path = self._ensure_correct_extension(file_path)

        # Se parsea ANTES de crear el origen: si el archivo es ilegible no debe
        # quedar una fila huérfana sin movimientos en el listado.
        ext = os.path.splitext(file_path)[1].lower()
        if ext == ".xlsx":
            try:
                filas_materia, filas_corte = self._read_xlsx(file_path)
            except Exception as e:
                raise ValueError(f"No se pudo leer el archivo XLSX: {e}")
        else:
            try:
                filas_materia, filas_corte = self._read_xls(file_path)
            except Exception as e:
                raise ValueError(f"No se pudo leer el archivo XLS: {e}")

        # Un archivo que solo trajera hojas de corte igual es válido.
        if not filas_materia and not filas_corte:
            raise ValueError(
                "El archivo no contiene movimientos reconocibles. "
                "Verifique que las columnas tengan los encabezados esperados."
            )

        origen = EstadoDiarioOrigen(
            usuario_carga_id=usuario_id,
            tipo=EstadoDiarioOrigen.TIPO_ESTADO_DIARIO,
            rut=rut,
            fecha=fecha,
            nombre_archivo=nombre_archivo or os.path.basename(file_path),
            fecha_carga=datetime.now(timezone.utc),
        )
        self.db.add(origen)
        self.db.flush()  # asigna origen.id sin cerrar la transacción

        count = 0
        for row in filas_materia:
            ed = EstadoDiario(
                estado_diario_origen_id=origen.id,
                rol=row.get("rol"),
                rol_unico=row.get("rol_unico"),
                fecha_ingreso=row.get("fecha_ingreso"),
                caratulado=row.get("caratulado"),
                tribunal=row.get("tribunal"),
                tipo_causa=row.get("tipo_causa"),
                ubicacion=row.get("ubicacion"),
                fecha_ubicacion=row.get("fecha_ubicacion"),
                estado=row.get("estado"),
                corte=row.get("corte"),
            )

            # Try to match jurisdiccion from corte
            if ed.corte:
                jur = self.jurisdiccion_repo.find_by_nombre(ed.corte)
                if jur:
                    ed.jurisdiccion_id = jur.id

            self.db.add(ed)
            count += 1

        # Las causas de corte van a su propia tabla: tienen otras columnas y se
        # muestran en otra pantalla (submenú Corte).
        cortes = 0
        for row in filas_corte:
            registro = EstadoDiarioCorte(
                estado_diario_origen_id=origen.id,
                tipo=row["tipo"],
                numero_ingreso=row.get("numero_ingreso"),
                fecha_ingreso=row.get("fecha_ingreso"),
                caratulado=row.get("caratulado"),
                ubicacion=row.get("ubicacion"),
                fecha_ubicacion=row.get("fecha_ubicacion"),
                corte=row.get("corte"),
                tipo_recurso=row.get("tipo_recurso"),
                jurisdiccion_id=self._jurisdiccion_de_corte(row["tipo"]),
            )
            self.db.add(registro)
            cortes += 1

        self.db.commit()
        logger.info(
            "Importados %d movimientos y %d causas de corte para origen %d",
            count, cortes, origen.id,
        )
        return {
            "origen_id": origen.id,
            "movimientos_importados": count,
            "cortes_importados": cortes,
        }

    def _jurisdiccion_de_corte(self, tipo: str) -> Optional[int]:
        """Jurisdicción que corresponde a una hoja de corte.

        Se busca por nombre normalizado y no por el literal de la hoja porque
        no coinciden: la hoja dice "Corte Apelaciones" y la jurisdicción
        sembrada es "Corte de Apelaciones". Sin esto, las causas de corte
        quedaban sin jurisdicción y el permiso de visibilidad del estudio no
        podía acotarlas.
        """
        buscado = "corte suprema" if tipo == EstadoDiarioCorte.TIPO_SUPREMA else "corte de apelaciones"
        for jur in self.jurisdiccion_repo.find_all():
            if normalizar_texto(jur.nombre) == buscado:
                return jur.id
        return None

    @staticmethod
    def _ensure_correct_extension(file_path: str) -> str:
        """Si el archivo tiene extensión .xls pero es realmente XLSX (ZIP),
        lo renombra a .xlsx para que openpyxl lo acepte."""
        ext = os.path.splitext(file_path)[1].lower()
        with open(file_path, "rb") as f:
            header = f.read(4)

        is_zip = header[:2] == b"PK"
        is_ole2 = header[:4] == b"\xd0\xcf\x11\xe0"

        if is_zip and ext != ".xlsx":
            new_path = file_path.rsplit(".", 1)[0] + ".xlsx"
            shutil.move(file_path, new_path)
            return new_path
        if is_ole2 and ext != ".xls":
            new_path = file_path.rsplit(".", 1)[0] + ".xls"
            shutil.move(file_path, new_path)
            return new_path
        return file_path

    def _read_xls(self, file_path: str) -> tuple[list[dict], list[dict]]:
        """Devuelve (filas de materia, filas de corte)."""
        wb = xlrd.open_workbook(file_path)
        materia: list[dict] = []
        cortes: list[dict] = []

        for sheet_idx in range(wb.nsheets):
            ws = wb.sheet_by_index(sheet_idx)
            if ws.nrows < 2:
                continue

            tipo_corte = tipo_de_hoja_corte(ws.name)
            alias = HEADER_ALIASES_CORTE if tipo_corte else HEADER_ALIASES

            col_map = {}
            for col_idx in range(ws.ncols):
                campo = self._campo_de_encabezado(ws.cell_value(0, col_idx), alias)
                if campo:
                    col_map[col_idx] = campo

            if not col_map:
                continue

            for row_idx in range(1, ws.nrows):
                row_data = {}
                for col_idx, field in col_map.items():
                    value = ws.cell_value(row_idx, col_idx)
                    if field in ("fecha_ingreso", "fecha_ubicacion"):
                        value = self._parse_date_xls(value, wb.datemode)
                    else:
                        value = str(value).strip() if value else None
                    row_data[field] = value

                if tipo_corte:
                    if any(row_data.values()):
                        row_data["tipo"] = tipo_corte
                        cortes.append(row_data)
                    continue

                if not row_data.get("corte"):
                    row_data["corte"] = ws.name
                if any(v for k, v in row_data.items() if k != "corte"):
                    materia.append(row_data)

        return materia, cortes

    @staticmethod
    def _campo_de_encabezado(encabezado, alias: dict) -> Optional[str]:
        """Campo interno que corresponde a un encabezado del Excel.

        La comparación va normalizada (sin acentos, sin grados, en minúsculas)
        porque el mismo dato viene escrito distinto en cada archivo: "N°
        Ingreso", "Número de Ingreso", "Ubicación"/"Ubicacion".
        """
        return alias.get(normalizar_texto(encabezado))

    def _read_xlsx(self, file_path: str) -> tuple[list[dict], list[dict]]:
        """Devuelve (filas de materia, filas de corte)."""
        wb = openpyxl.load_workbook(file_path)
        materia: list[dict] = []
        cortes: list[dict] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row is None or ws.max_row < 2:
                continue

            tipo_corte = tipo_de_hoja_corte(sheet_name)
            col_map = self._build_col_map(
                ws, HEADER_ALIASES_CORTE if tipo_corte else HEADER_ALIASES
            )
            if not col_map:
                continue

            for row_idx in range(2, ws.max_row + 1):
                row_data = {}
                for col_idx, field in col_map.items():
                    value = ws.cell(row_idx, col_idx).value
                    if field in ("fecha_ingreso", "fecha_ubicacion"):
                        value = self._parse_date_xlsx(value)
                    else:
                        value = str(value).strip() if value else None
                    row_data[field] = value

                if tipo_corte:
                    if any(row_data.values()):
                        row_data["tipo"] = tipo_corte
                        cortes.append(row_data)
                    continue

                # Use sheet name as corte if not already set
                if not row_data.get("corte"):
                    row_data["corte"] = sheet_name

                if any(v for k, v in row_data.items() if k != "corte"):
                    materia.append(row_data)
        wb.close()
        return materia, cortes

    @classmethod
    def _build_col_map(cls, ws, alias: dict) -> dict[int, str]:
        """Índice de columna (base 1) -> campo interno, según los encabezados."""
        col_map = {}
        for col_idx in range(1, (ws.max_column or 0) + 1):
            campo = cls._campo_de_encabezado(ws.cell(1, col_idx).value, alias)
            if campo:
                col_map[col_idx] = campo
        return col_map

    def _parse_date_xls(self, value, datemode) -> Optional[date]:
        if not value:
            return None
        try:
            if isinstance(value, float):
                dt = xlrd.xldate_as_datetime(value, datemode)
                return dt.date()
            return datetime.strptime(str(value).strip(), "%d/%m/%Y").date()
        except Exception:
            try:
                return datetime.strptime(str(value).strip(), "%Y-%m-%d").date()
            except Exception:
                return None

    def _parse_date_xlsx(self, value) -> Optional[date]:
        if not value:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        try:
            return datetime.strptime(str(value).strip(), "%d/%m/%Y").date()
        except Exception:
            try:
                return datetime.strptime(str(value).strip(), "%Y-%m-%d").date()
            except Exception:
                return None
