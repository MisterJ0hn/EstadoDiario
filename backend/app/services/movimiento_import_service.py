"""Importación del Excel de "Movimientos".

Es un reporte distinto al estado diario: lista el universo de causas vigentes
de un RUT con su estado procesal. Diferencias que condicionan este parser:

* El archivo real es un .xls OLE2 (el del estado diario es un XLSX con
  extensión .xls mentida), así que el formato se detecta por magic bytes y no
  por la extensión.
* Trae una hoja por materia y la materia NO viene como columna: es el nombre
  de la hoja.
* El orden de las columnas cambia entre hojas (en Familia, "Fecha Ingreso" va
  antes que "Caratulado") y los encabezados varían en tildes y mayúsculas
  ("Institución"/"Institucion", "Estado Causa"/"Estado causa"). Por eso el
  mapeo es por nombre normalizado de encabezado, nunca por posición.
* Varias hojas vienen solo con el encabezado; se saltan sin error.
"""

import logging
import os
import re
from datetime import datetime, timezone, date
from typing import Optional

import openpyxl
import xlrd
from sqlalchemy.orm import Session

from app.models.estado_diario_origen import EstadoDiarioOrigen
from app.models.movimiento import Movimiento
from app.models.movimiento_corte import MovimientoCorte
from app.repositories.movimiento_repository import MovimientoRepository
from app.repositories.jurisdiccion_repository import JurisdiccionRepository
from app.services.deteccion_archivo import verificar_contenido
from app.utils.excel_pjud import (
    detectar_formato,
    mapa_columnas,
    parse_fecha_xls,
    parse_fecha_xlsx,
    recortar,
)

from app.services.cartera_sync_service import sincronizar_cartera
from app.services.import_service import normalizar_texto, tipo_de_hoja_corte

logger = logging.getLogger(__name__)

# Encabezado normalizado (minúsculas, sin tildes, sin espacios repetidos) ->
# campo del modelo Movimiento. "Rit" en primera instancia y "Rol" en las
# cortes son la misma columna lógica.
HEADER_ALIASES = {
    "rit": "rol",
    "rol": "rol",
    "era": "era",
    "tribunal": "tribunal",
    "corte": "corte",
    "caratulado": "caratulado",
    "fecha ingreso": "fecha_ingreso",
    "fecha de ingreso": "fecha_ingreso",
    "estado causa": "estado_causa",
    "estado de causa": "estado_causa",
    "estado": "estado_causa",
    "institucion": "institucion",
    "ubicacion": "ubicacion",
    "fecha ubicacion": "fecha_ubicacion",
}

CAMPOS_FECHA = ("fecha_ingreso", "fecha_ubicacion")

# Largos de las columnas del modelo. El Excel del PJUD a veces trae textos
# rellenos con espacios o más largos de lo esperado: se recortan en vez de
# reventar el INSERT.
LARGOS_MAXIMOS = {
    "materia": 100,
    "rol": 100,
    "era": 20,
    "tribunal": 255,
    "corte": 255,
    "caratulado": 255,
    "estado_causa": 100,
    "institucion": 255,
    "ubicacion": 255,
}


def _recortar(valor: Optional[str], campo: str) -> Optional[str]:
    return recortar(valor, campo, LARGOS_MAXIMOS)


def _mapa_columnas(encabezados: list) -> dict:
    return mapa_columnas(encabezados, HEADER_ALIASES)


def _armar_fila(valores_por_campo: dict, nombre_hoja: str) -> Optional[dict]:
    """Devuelve la fila lista para persistir, o None si venía vacía.

    Las hojas de corte se marcan con `tipo`: van a `movimiento_corte`, no a
    `movimiento`. Traen columnas que las hojas de materia no tienen (Era,
    Ubicación, Fecha Ubicación) y les falta el tribunal, así que mezclarlas
    dejaba media tabla vacía en las dos direcciones.
    """
    if not any(v not in (None, "") for v in valores_por_campo.values()):
        return None

    tipo_corte = tipo_de_hoja_corte(nombre_hoja)
    if tipo_corte:
        fila = {"tipo": tipo_corte}
    else:
        # En las hojas de materia, la materia ES el nombre de la hoja: no
        # viene como columna.
        fila = {"materia": _recortar(nombre_hoja.strip(), "materia")}

    fila.update(valores_por_campo)
    return fila


def _leer_xls(file_path: str) -> list[dict]:
    wb = xlrd.open_workbook(file_path)
    filas: list[dict] = []
    for hoja in wb.sheets():
        if hoja.nrows < 2:  # solo encabezado (o vacía)
            continue

        mapa = _mapa_columnas([hoja.cell_value(0, c) for c in range(hoja.ncols)])
        if not mapa:
            continue

        for r in range(1, hoja.nrows):
            valores = {}
            for col, campo in mapa.items():
                bruto = hoja.cell_value(r, col)
                if campo in CAMPOS_FECHA:
                    valores[campo] = parse_fecha_xls(bruto, wb.datemode)
                else:
                    texto = str(bruto).strip() if bruto not in (None, "") else None
                    valores[campo] = _recortar(texto or None, campo)

            fila = _armar_fila(valores, hoja.name)
            if fila:
                filas.append(fila)
    return filas


def _leer_xlsx(file_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    try:
        filas: list[dict] = []
        for nombre_hoja in wb.sheetnames:
            hoja = wb[nombre_hoja]
            if not hoja.max_row or hoja.max_row < 2:
                continue

            ncols = hoja.max_column or 0
            mapa = _mapa_columnas([hoja.cell(1, c).value for c in range(1, ncols + 1)])
            if not mapa:
                continue

            for r in range(2, hoja.max_row + 1):
                valores = {}
                for col, campo in mapa.items():
                    bruto = hoja.cell(r, col + 1).value  # openpyxl es 1-based
                    if campo in CAMPOS_FECHA:
                        valores[campo] = parse_fecha_xlsx(bruto)
                    else:
                        texto = str(bruto).strip() if bruto not in (None, "") else None
                        valores[campo] = _recortar(texto or None, campo)

                fila = _armar_fila(valores, nombre_hoja)
                if fila:
                    filas.append(fila)
        return filas
    finally:
        wb.close()


def parse_movimientos_file(file_path: str) -> list[dict]:
    """Parsea el Excel de movimientos a una lista de dicts.

    Función pura (no toca la base de datos) para poder probar el parseo del
    archivo por separado.
    """
    if detectar_formato(file_path) == "xlsx":
        return _leer_xlsx(file_path)
    return _leer_xls(file_path)


def parse_nombre_archivo(filename: str) -> tuple[Optional[str], Optional[date]]:
    """Extrae RUT y fecha del nombre, formato Movimientos_{RUT}_{DD}_{MM}_{YYYY}.xls.

    El archivo real trae doble guion bajo cuando el RUT viene sin dígito
    verificador (Movimientos_16952077__30_07_2026.xls), de ahí el `_+`.
    """
    nombre = os.path.splitext(filename or "")[0]
    m = re.match(
        r"^Movimientos_(\d+(?:[\-]?[0-9kK])?)_+(\d{1,2})_(\d{1,2})_(\d{4})$",
        nombre,
        re.IGNORECASE,
    )
    if not m:
        return None, None
    try:
        return m.group(1), date(int(m.group(4)), int(m.group(3)), int(m.group(2)))
    except ValueError:
        return m.group(1), None


class MovimientoImportService:
    # La fecha de este reporte es la de su emisión y no aparece en ninguna
    # columna: las fechas de adentro son de ingreso de cada causa, que es otra
    # cosa. Si el nombre no la trae, no hay de dónde deducirla.
    deduce_fecha_del_contenido = False
    # Igual que el estado diario: su control de duplicados es (rut, fecha).
    requiere_rut = True

    def __init__(self, db: Session):
        self.db = db
        self.repo = MovimientoRepository(db)
        self.jurisdiccion_repo = JurisdiccionRepository(db)

    def _jurisdiccion_de_corte(self, tipo: str) -> Optional[int]:
        """Jurisdicción que corresponde a una hoja de corte.

        Se busca por nombre normalizado porque la hoja dice "Corte
        Apelaciones" y la jurisdicción sembrada es "Corte de Apelaciones". Sin
        esto, las causas de corte quedan sin jurisdicción y el permiso de
        visibilidad del estudio no puede acotarlas.
        """
        buscado = "corte suprema" if tipo == MovimientoCorte.TIPO_SUPREMA else "corte de apelaciones"
        for jur in self.jurisdiccion_repo.find_all():
            if normalizar_texto(jur.nombre) == buscado:
                return jur.id
        return None

    def import_file(
        self,
        file_path: str,
        rut: str,
        fecha: date,
        usuario_id: Optional[int] = None,
        nombre_archivo: Optional[str] = None,
    ) -> dict:
        """Importa un archivo de movimientos y devuelve el resumen de la carga.

        Mismo contrato que ImportService.import_file: rechaza duplicados y
        parsea ANTES de crear el origen, para que un archivo ilegible no deje
        una fila huérfana sin movimientos en el listado de archivos.
        """
        # Que el archivo sea de verdad de movimientos y no otro reporte mal
        # ruteado (ver `deteccion_archivo`).
        error = verificar_contenido(file_path, EstadoDiarioOrigen.TIPO_MOVIMIENTOS)
        if error:
            raise ValueError(error)

        # El duplicado se mide dentro del tipo: un RUT puede tener el mismo día
        # un estado diario y un archivo de movimientos, pero no dos de estos.
        existente = self.repo.find_origen_movimientos(rut, fecha)
        if existente:
            raise ValueError(
                f"Ya existe un archivo de movimientos para el RUT {rut} del {fecha} "
                f"(origen {existente.id}). Elimínelo antes de volver a cargarlo."
            )

        try:
            filas = parse_movimientos_file(file_path)
        except Exception as e:
            raise ValueError(f"No se pudo leer el archivo de movimientos: {e}")

        if not filas:
            raise ValueError(
                "El archivo no contiene movimientos reconocibles. Verifique que "
                "las hojas tengan los encabezados esperados (Rit/Rol, Tribunal, "
                "Caratulado, Fecha Ingreso, Estado Causa, Institución)."
            )

        origen = EstadoDiarioOrigen(
            usuario_carga_id=usuario_id,
            tipo=EstadoDiarioOrigen.TIPO_MOVIMIENTOS,
            rut=rut,
            fecha=fecha,
            nombre_archivo=nombre_archivo or os.path.basename(file_path),
            fecha_carga=datetime.now(timezone.utc),
        )
        self.db.add(origen)
        self.db.flush()  # asigna origen.id sin cerrar la transacción

        cache_jurisdiccion: dict[str, Optional[int]] = {}
        por_materia: dict[str, int] = {}

        cortes = 0
        for fila in filas:
            if fila.get("tipo"):
                self.db.add(
                    MovimientoCorte(
                        estado_diario_origen_id=origen.id,
                        tipo=fila["tipo"],
                        rol=fila.get("rol"),
                        era=fila.get("era"),
                        corte=fila.get("corte"),
                        fecha_ingreso=fila.get("fecha_ingreso"),
                        ubicacion=fila.get("ubicacion"),
                        fecha_ubicacion=fila.get("fecha_ubicacion"),
                        caratulado=fila.get("caratulado"),
                        estado_causa=fila.get("estado_causa"),
                        institucion=fila.get("institucion"),
                        jurisdiccion_id=self._jurisdiccion_de_corte(fila["tipo"]),
                    )
                )
                cortes += 1
                continue

            mov = Movimiento(
                estado_diario_origen_id=origen.id,
                materia=fila.get("materia"),
                rol=fila.get("rol"),
                era=fila.get("era"),
                tribunal=fila.get("tribunal"),
                corte=fila.get("corte"),
                caratulado=fila.get("caratulado"),
                fecha_ingreso=fila.get("fecha_ingreso"),
                estado_causa=fila.get("estado_causa"),
                institucion=fila.get("institucion"),
                ubicacion=fila.get("ubicacion"),
                fecha_ubicacion=fila.get("fecha_ubicacion"),
            )

            # Solo la hoja "Corte Apelaciones" trae la columna Corte; es el
            # único dato que permite amarrar la fila a una jurisdicción.
            if mov.corte:
                if mov.corte not in cache_jurisdiccion:
                    jur = self.jurisdiccion_repo.find_by_nombre(mov.corte)
                    cache_jurisdiccion[mov.corte] = jur.id if jur else None
                mov.jurisdiccion_id = cache_jurisdiccion[mov.corte]

            self.db.add(mov)
            por_materia[mov.materia or "(sin materia)"] = (
                por_materia.get(mov.materia or "(sin materia)", 0) + 1
            )

        # El cruce va ANTES del commit y en la misma transacción: si fallara,
        # no queda un archivo importado y una cartera a medio actualizar.
        self.db.flush()
        sincronizar_cartera(self.db)

        self.db.commit()
        # `len(filas)` incluiría las de corte, que van a otra tabla: el número
        # que se informa tiene que ser el de filas realmente insertadas acá.
        movimientos = sum(por_materia.values())
        logger.info(
            "Importados %d movimientos y %d causas de corte para origen %d (%s)",
            movimientos, cortes, origen.id, por_materia,
        )
        return {
            "origen_id": origen.id,
            "movimientos_importados": movimientos,
            "cortes_importados": cortes,
            "por_materia": por_materia,
        }
