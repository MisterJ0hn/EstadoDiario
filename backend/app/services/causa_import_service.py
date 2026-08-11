"""Importación del Excel de "Causas".

Es el tercer reporte del PJUD que entra al sistema y hay que no confundirlo con
los otros dos:

* **Estado diario**: lo que se movió UN día.
* **Movimientos**: el estado procesal de las causas vigentes.
* **Causas** (éste): la **cartera** completa del estudio — todas las causas en
  que es parte, se hayan movido o no.

Particularidades del archivo, todas verificadas contra
`ejemplos/Causas_16952077-1.xlsx`:

* Las hojas de materia **no traen las mismas columnas**: Cobranza no tiene
  Estado Causa, Penal agrega Tipo Causa y Ruc, Familia cambia el orden. Por eso
  el mapeo es por encabezado normalizado y nunca por posición.
* El identificador se llama "Rol" en Civil/Laboral/Cobranza y "Rit" en
  Penal/Familia: es el mismo dato y va al campo `rol`.
* Las dos hojas de corte nombran distinto la misma columna: Suprema dice
  "Estado Causa" y Apelaciones "Estado Procesal". Las dos van a
  `estado_procesal`.
* La materia no viene como columna: es el nombre de la hoja.
"""

import logging
import os
import re
from datetime import date, datetime, timezone
from typing import Optional

import openpyxl
import xlrd
from sqlalchemy.orm import Session

from app.core.estados_causa import esta_finalizada
from app.models.causa import Causa
from app.models.causa_corte import CausaCorte
from app.models.estado_diario_origen import EstadoDiarioOrigen
from app.repositories.causa_repository import CausaRepository
from app.repositories.jurisdiccion_repository import JurisdiccionRepository
from app.services.cartera_sync_service import sincronizar_cartera
from app.services.import_service import normalizar_texto, tipo_de_hoja_corte
from app.utils.excel_pjud import (
    detectar_formato,
    parse_fecha_xls,
    parse_fecha_xlsx,
)

logger = logging.getLogger(__name__)

# Encabezado normalizado -> campo del modelo, para las hojas de MATERIA.
HEADER_ALIASES = {
    "tipo causa": "tipo_causa",
    "tipo de causa": "tipo_causa",
    "rol": "rol",
    "rit": "rol",
    "rol rit": "rol",
    "ruc": "ruc",
    "rol unico": "ruc",
    "tribunal": "tribunal",
    "fecha ingreso": "fecha_ingreso",
    "fecha de ingreso": "fecha_ingreso",
    "caratulado": "caratulado",
    "estado causa": "estado_causa",
    "estado de causa": "estado_causa",
    "estado": "estado_causa",
    "institucion": "institucion",
}

# Encabezados de las hojas de CORTE. Van aparte porque chocan con los de
# materia: acá "estado causa" es el estado PROCESAL del recurso, no el de una
# causa de primera instancia, y se guarda en otra columna.
HEADER_ALIASES_CORTE = {
    "rol": "rol",
    "era": "era",
    "corte": "corte",
    "fecha ingreso": "fecha_ingreso",
    "fecha de ingreso": "fecha_ingreso",
    "ubicacion": "ubicacion",
    "fecha ubicacion": "fecha_ubicacion",
    "fecha de ubicacion": "fecha_ubicacion",
    "caratulado": "caratulado",
    # Las dos hojas la nombran distinto y es el mismo dato.
    "estado procesal": "estado_procesal",
    "estado causa": "estado_procesal",
    "estado de causa": "estado_procesal",
    "institucion": "institucion",
}

CAMPOS_FECHA = {"fecha_ingreso", "fecha_ubicacion"}

# Largo de cada columna en el modelo. Un caratulado de 700 caracteres no puede
# tumbar la importación completa con un DataError: se recorta y se sigue.
LARGOS = {
    "materia": 100,
    "tipo_causa": 100,
    "rol": 100,
    "ruc": 100,
    "era": 20,
    "tribunal": 255,
    "caratulado": 500,
    "estado_causa": 255,
    "estado_procesal": 255,
    "institucion": 255,
    "corte": 255,
    "ubicacion": 255,
}


def _recortar(valor: Optional[str], campo: str) -> Optional[str]:
    tope = LARGOS.get(campo)
    if valor is None or tope is None:
        return valor
    return valor[:tope]


def _mapa_columnas(encabezados: list, es_corte: bool) -> dict[int, str]:
    """{índice de columna -> campo}. Vacío si la hoja no es reconocible."""
    alias = HEADER_ALIASES_CORTE if es_corte else HEADER_ALIASES
    mapa = {}
    for i, bruto in enumerate(encabezados):
        campo = alias.get(normalizar_texto(bruto or ""))
        # El primero gana: si una hoja repitiera un encabezado, quedarse con el
        # último dejaría la columna buena pisada por una vacía.
        if campo and campo not in mapa.values():
            mapa[i] = campo
    return mapa


def _armar_fila(valores: dict, nombre_hoja: str, tipo_corte: Optional[str]) -> Optional[dict]:
    """Fila lista para persistir, o None si venía en blanco."""
    if not any(v not in (None, "") for v in valores.values()):
        return None

    if tipo_corte:
        fila = {"tipo": tipo_corte}
    else:
        # En las hojas de materia, la materia ES el nombre de la hoja.
        fila = {"materia": _recortar(nombre_hoja.strip(), "materia")}

    fila.update(valores)
    return fila


def _leer_xls(file_path: str) -> list[dict]:
    wb = xlrd.open_workbook(file_path)
    filas: list[dict] = []
    for hoja in wb.sheets():
        if hoja.nrows < 2:  # solo encabezado, o vacía
            continue

        tipo_corte = tipo_de_hoja_corte(hoja.name)
        mapa = _mapa_columnas(
            [hoja.cell_value(0, c) for c in range(hoja.ncols)], bool(tipo_corte)
        )
        if not mapa:
            continue

        for r in range(1, hoja.nrows):
            valores = {}
            for col, campo in mapa.items():
                bruto = hoja.cell_value(r, col)
                if campo in CAMPOS_FECHA:
                    valores[campo] = parse_fecha_xls(bruto, wb.datemode)
                else:
                    texto = str(bruto).strip() if bruto not in (None, "") else None
                    valores[campo] = _recortar(texto or None, campo)

            fila = _armar_fila(valores, hoja.name, tipo_corte)
            if fila:
                filas.append(fila)
    return filas


def _leer_xlsx(file_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    try:
        filas: list[dict] = []
        for nombre_hoja in wb.sheetnames:
            hoja = wb[nombre_hoja]
            tipo_corte = tipo_de_hoja_corte(nombre_hoja)

            # read_only no permite acceso aleatorio por celda: se recorre en
            # streaming. Con 9.000 filas en Civil, cargar la hoja completa en
            # memoria por archivo es justamente lo que hay que evitar.
            mapa: dict[int, str] = {}
            for i, fila_bruta in enumerate(hoja.iter_rows(values_only=True)):
                if i == 0:
                    mapa = _mapa_columnas(list(fila_bruta), bool(tipo_corte))
                    if not mapa:
                        break
                    continue

                valores = {}
                for col, campo in mapa.items():
                    bruto = fila_bruta[col] if col < len(fila_bruta) else None
                    if campo in CAMPOS_FECHA:
                        valores[campo] = parse_fecha_xlsx(bruto)
                    else:
                        texto = str(bruto).strip() if bruto not in (None, "") else None
                        valores[campo] = _recortar(texto or None, campo)

                fila = _armar_fila(valores, nombre_hoja, tipo_corte)
                if fila:
                    filas.append(fila)
        return filas
    finally:
        wb.close()


def parse_causas_file(file_path: str) -> list[dict]:
    """Parsea el Excel de causas a una lista de dicts.

    Función pura (no toca la base) para poder probar el parseo por separado.
    """
    if detectar_formato(file_path) == "xlsx":
        return _leer_xlsx(file_path)
    return _leer_xls(file_path)


def parse_nombre_archivo(filename: str) -> tuple[Optional[str], Optional[date]]:
    """Extrae el RUT del nombre: `Causas_{RUT}.xlsx`.

    **Este reporte no trae fecha en el nombre**, a diferencia de movimientos y
    audiencias: es una foto de la cartera al momento de emitirlo. La fecha la
    pone quien lo carga (por defecto, hoy).
    """
    nombre = os.path.splitext(filename or "")[0]
    m = re.match(r"^Causas_(\d+(?:[\-]?[0-9kK])?)_*$", nombre, re.IGNORECASE)
    return (m.group(1) if m else None), None


class CausaImportService:
    # La fecha no está en el archivo ni en su nombre: es la de emisión del
    # reporte, y ninguna columna la trae (las de adentro son de ingreso de
    # cada causa, que es otra cosa).
    deduce_fecha_del_contenido = False
    # Mismo control de duplicados que los otros reportes: (rut, fecha, tipo).
    requiere_rut = True

    def __init__(self, db: Session):
        self.db = db
        self.repo = CausaRepository(db)
        self.jurisdiccion_repo = JurisdiccionRepository(db)

    # ── Deduplicación ─────────────────────────────────────
    #
    # El Excel del PJUD trae la misma causa más de una vez. En una cartera real
    # de 12.248 filas había 1.653 repetidas, y en cero de esos grupos difería el
    # caratulado: es la misma causa, no dos parecidas.
    #
    # **La llave es `rol + tribunal`**, medido sobre esa misma cartera: el rol
    # se renumera en cada tribunal (`C-17-2021` existe en ocho juzgados civiles
    # a la vez), y agregarle la materia da exactamente el mismo número de
    # causas, porque el tribunal ya determina la materia.

    @staticmethod
    def _norm(valor: Optional[str]) -> str:
        """Para comparar: sin espacios sobrantes y en minúsculas. El Excel
        rellena las celdas a la derecha ('Concluido            ')."""
        return (valor or "").strip().lower()

    @classmethod
    def _clave_materia(cls, fila: dict) -> Optional[tuple]:
        """`(rol, tribunal)`, o None si falta alguna de las dos.

        Sin las dos partes no se puede afirmar que dos filas sean la misma
        causa, así que entran las dos por separado: preferible una repetida a
        fundir dos causas distintas, que sería perder una.
        """
        rol, tribunal = cls._norm(fila.get("rol")), cls._norm(fila.get("tribunal"))
        return (rol, tribunal) if rol and tribunal else None

    @classmethod
    def _clave_corte(cls, fila: dict) -> Optional[tuple]:
        """`(tipo, rol, era, corte)`. En Suprema `corte` viene vacío —hay una
        sola— así que ahí la llave efectiva es rol + era."""
        rol = cls._norm(fila.get("rol"))
        if not rol:
            return None
        return (cls._norm(fila.get("tipo")), rol,
                cls._norm(fila.get("era")), cls._norm(fila.get("corte")))

    @classmethod
    def _fusionar_materia(cls, causa: Causa, fila: dict) -> None:
        """Vuelca sobre la causa ya vista lo que aporte su duplicado.

        **Si alguna de las dos dice que terminó, la causa terminó.** Es la regla
        que pidió el estudio y es la conservadora: una causa concluida no se
        factura, así que ante la duda no se cobra. Sin esto quedaba el estado de
        la fila que viniera primero en el Excel, que es un orden que nadie
        controla.
        """
        if esta_finalizada(fila.get("estado_causa")):
            causa.estado_causa = fila.get("estado_causa")
        elif not (causa.estado_causa or "").strip():
            causa.estado_causa = fila.get("estado_causa")

        # El resto solo rellena huecos: dos filas de la misma causa pueden traer
        # una el RUC y la otra no.
        for campo in ("tipo_causa", "ruc", "fecha_ingreso", "caratulado", "institucion"):
            if not getattr(causa, campo, None) and fila.get(campo):
                setattr(causa, campo, fila[campo])

    @classmethod
    def _fusionar_corte(cls, causa: CausaCorte, fila: dict) -> None:
        """Igual que en materia, sobre `estado_procesal`, que es como se llama
        ahí la misma columna."""
        if esta_finalizada(fila.get("estado_procesal")):
            causa.estado_procesal = fila.get("estado_procesal")
        elif not (causa.estado_procesal or "").strip():
            causa.estado_procesal = fila.get("estado_procesal")

        for campo in ("era", "corte", "fecha_ingreso", "ubicacion",
                      "fecha_ubicacion", "caratulado", "institucion"):
            if not getattr(causa, campo, None) and fila.get(campo):
                setattr(causa, campo, fila[campo])

    def _jurisdiccion_de_corte(self, tipo: str) -> Optional[int]:
        """Jurisdicción de una hoja de corte.

        Por nombre normalizado porque la hoja dice "Corte Apelaciones" y la
        jurisdicción sembrada es "Corte de Apelaciones".
        """
        buscado = "corte suprema" if tipo == CausaCorte.TIPO_SUPREMA else "corte de apelaciones"
        for jur in self.jurisdiccion_repo.find_all():
            if normalizar_texto(jur.nombre) == buscado:
                return jur.id
        return None

    def import_file(
        self,
        file_path: str,
        rut: str,
        fecha: date,
        usuario_id: Optional[int] = None,
        nombre_archivo: Optional[str] = None,
    ) -> dict:
        """Importa el archivo de causas y devuelve el resumen de la carga."""
        existente = self.repo.find_origen(rut, fecha)
        if existente:
            raise ValueError(
                f"Ya existe un archivo de causas para el RUT {rut} del {fecha} "
                f"(origen {existente.id}). Elimínelo antes de volver a cargarlo."
            )

        try:
            filas = parse_causas_file(file_path)
        except Exception as e:
            raise ValueError(f"No se pudo leer el archivo de causas: {e}")

        if not filas:
            raise ValueError(
                "El archivo no contiene causas reconocibles. Verifique que las "
                "hojas tengan los encabezados esperados (Rol/Rit, Tribunal, "
                "Caratulado, Fecha Ingreso, Institución)."
            )

        # El origen se crea DESPUÉS de parsear: si el archivo no sirve, no
        # queda una fila fantasma en Bitácora.
        origen = EstadoDiarioOrigen(
            usuario_carga_id=usuario_id,
            tipo=EstadoDiarioOrigen.TIPO_CAUSAS,
            rut=rut,
            fecha=fecha,
            nombre_archivo=nombre_archivo or os.path.basename(file_path),
            fecha_carga=datetime.now(timezone.utc),
        )
        self.db.add(origen)
        self.db.flush()  # asigna origen.id sin cerrar la transacción

        cache_jurisdiccion: dict[str, Optional[int]] = {}
        por_materia: dict[str, int] = {}
        cortes = 0
        # Las causas ya vistas en ESTE archivo, por su llave. El Excel del PJUD
        # repite la misma causa —1.653 filas de más en una cartera real de
        # 12.248— y sin esto entraban todas: la pantalla las mostraba dos veces
        # y la factura las cobraba dos veces.
        vistas_materia: dict[tuple, Causa] = {}
        vistas_corte: dict[tuple, CausaCorte] = {}
        duplicadas = 0

        for fila in filas:
            if fila.get("tipo"):
                clave = self._clave_corte(fila)
                previa = vistas_corte.get(clave) if clave else None
                if previa is not None:
                    self._fusionar_corte(previa, fila)
                    duplicadas += 1
                    continue

                registro = CausaCorte(
                    estado_diario_origen_id=origen.id,
                    tipo=fila["tipo"],
                    rol=fila.get("rol"),
                    era=fila.get("era"),
                    corte=fila.get("corte"),
                    fecha_ingreso=fila.get("fecha_ingreso"),
                    ubicacion=fila.get("ubicacion"),
                    fecha_ubicacion=fila.get("fecha_ubicacion"),
                    caratulado=fila.get("caratulado"),
                    estado_procesal=fila.get("estado_procesal"),
                    institucion=fila.get("institucion"),
                    jurisdiccion_id=self._jurisdiccion_de_corte(fila["tipo"]),
                )
                self.db.add(registro)
                if clave:
                    vistas_corte[clave] = registro
                cortes += 1
                continue

            materia = fila.get("materia")
            if materia not in cache_jurisdiccion:
                jur = self.jurisdiccion_repo.find_by_nombre(materia) if materia else None
                cache_jurisdiccion[materia] = jur.id if jur else None

            clave = self._clave_materia(fila)
            previa = vistas_materia.get(clave) if clave else None
            if previa is not None:
                self._fusionar_materia(previa, fila)
                duplicadas += 1
                continue

            causa = Causa(
                estado_diario_origen_id=origen.id,
                materia=materia,
                tipo_causa=fila.get("tipo_causa"),
                rol=fila.get("rol"),
                ruc=fila.get("ruc"),
                tribunal=fila.get("tribunal"),
                fecha_ingreso=fila.get("fecha_ingreso"),
                caratulado=fila.get("caratulado"),
                estado_causa=fila.get("estado_causa"),
                institucion=fila.get("institucion"),
                jurisdiccion_id=cache_jurisdiccion[materia],
            )
            self.db.add(causa)
            if clave:
                vistas_materia[clave] = causa
            por_materia[materia] = por_materia.get(materia, 0) + 1

        # La carga de Causas reemplaza la foto de la cartera, así que el cruce
        # se rehace sobre el origen recién creado: lo enriquecido antes colgaba
        # del archivo anterior y ya nadie lo lee.
        self.db.flush()
        sincronizar_cartera(self.db)

        self.db.commit()

        causas = sum(por_materia.values())
        logger.info(
            "Importadas %d causas y %d de corte para origen %d (%s); "
            "%d filas repetidas del Excel se fusionaron",
            causas, cortes, origen.id, por_materia, duplicadas,
        )
        return {
            "origen_id": origen.id,
            "causas_importadas": causas,
            "cortes_importados": cortes,
            "duplicadas_fusionadas": duplicadas,
            "por_materia": por_materia,
        }
